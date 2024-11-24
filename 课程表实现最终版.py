import openpyxl
from openpyxl.styles import Alignment,Font

#创建一个水平和垂直都居中对齐对象
#设置自动换行
###得写在同一个alignment里面 不然只会执行最后一个关于alignment的操作
alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)

#创建一个字体对象，并设置字体、大小、颜色等属性
font_style = Font(name='微软雅黑', size=11, color='000000', bold=False, italic=False)

# 加载txt文件 并生成schedule的形式
def parse_txt_to_schedule(file_path):
    schedule = {}
    current_day = None
    with open(file_path, 'r', encoding='utf-8') as file:
        lines = file.readlines()
    for line in lines:
        line = line.strip() #过滤内容
        if line.startswith("星期"):
            current_day = line
            schedule[current_day] = []
        elif line:  #去除空字符！！！ GPT你怎么能这么聪明
            schedule[current_day].append(line)
    # 将时间、课程、地点分组
    structured_schedule = {}
    for day, details in schedule.items(): #键值对形式赋值（key,value）
        structured_schedule[day] = [] #初始化空列表
        for i in range(0, len(details), 3):  # 每三行是一个完整的条目
            time, course, location = details[i:i+3]
            structured_schedule[day].append((time, course, location))
    return structured_schedule

# 写入Excel文件
def write_excel(schedule, template_file, output_file):
    wb = openpyxl.load_workbook(template_file)
    ws = wb.active
    time_to_row = {
        "08:00": 2,
        "09:00": 3,
        "10:00": 4,
        "11:00": 5,
        "12:00": 6,
        "13:00": 7,
        "14:00": 8,
        "15:00": 9,
        "16:00": 10,
        "17:00": 11,
        "18:00": 12,
        "19:00": 13,
        "20:00": 14,
        "21:00": 15,
        "22:00": 16
    }
    day_to_col = {
        "星期一": 2,
        "星期二": 3,
        "星期三": 4,
        "星期四": 5,
        "星期五": 6,
        "星期六": 7,
        "星期天": 8,
    }
    for day, classes in schedule.items():
        col = day_to_col.get(day, None)
        for time, course, location in classes:
            # 将时间映射到所在的小时（如 09:50 -> 09:00）
            hour = f"{int(time.split(':')[0]):02d}:00"
            row = time_to_row.get(hour, None)
            # 合并课程名称、地点和时间到一个单元格，并按三行排列
            cell_value = f"{course}\n{location}\n{time}"
            ws.cell(row=row, column=col, value=cell_value)
            ws.cell(row=row, column=col).font=font_style
            ws.cell(row=row, column=col).alignment=alignment
    
    # 保存Excel文件
    wb.save(output_file)
    print(f"课程表已保存到 {output_file}")
# 主函数
if __name__ == "__main__":
    txt_file_path =r"C:\Users\admin\Desktop\courses.txt" # 替换为实际路径
    template_file_path = r"C:\Users\admin\Desktop\Agenda Agent.xlsx"  # 替换为你的模板路径
    output_excel_file = r"C:\Users\admin\Desktop\agenda.xlsx"
    schedule = parse_txt_to_schedule(txt_file_path)
    write_excel(schedule, template_file_path, output_excel_file)